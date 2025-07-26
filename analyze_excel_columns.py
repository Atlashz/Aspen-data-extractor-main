#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Column Analyzer for BFG-CO2H-HEX.xlsx
Analyzes columns I through N to understand heat exchanger data structure
"""

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import os
from pathlib import Path

def analyze_with_openpyxl():
    """
    Analyze Excel file using openpyxl (raw Excel access)
    """
    excel_file = "BFG-CO2H-HEX.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} not found in current directory")
        return None, None
        
    try:
        print(f"Reading Excel file with openpyxl: {excel_file}")
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        
        print(f"\n=== EXCEL FILE ANALYSIS (openpyxl) ===")
        print(f"File: {excel_file}")
        print(f"Active sheet: {ws.title}")
        print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Get header row (assuming first row contains headers)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(cell_value)
        
        print(f"\n=== ALL COLUMN HEADERS ===")
        for i, header in enumerate(headers):
            excel_col_letter = chr(65 + i) if i < 26 else chr(65 + i//26 - 1) + chr(65 + i%26)
            print(f"Column {i+1:2d} ({excel_col_letter:2s}): {header}")
        
        # Focus on columns I through N (columns 9-14 in 1-based Excel numbering)
        print(f"\n=== COLUMNS I THROUGH N ANALYSIS ===")
        target_columns = []
        target_positions = range(9, 15)  # Columns I-N (9-14 in 1-based)
        
        for excel_col_num in target_positions:
            if excel_col_num <= ws.max_column:
                excel_letter = chr(65 + excel_col_num - 1) if excel_col_num <= 26 else chr(65 + (excel_col_num-1)//26 - 1) + chr(65 + (excel_col_num-1)%26)
                header = ws.cell(row=1, column=excel_col_num).value
                target_columns.append((excel_col_num-1, excel_letter, header))
                
                print(f"\nColumn {excel_col_num:2d} ({excel_letter}): '{header}'")
                
                # Sample data from first 10 rows
                sample_values = []
                for row in range(2, min(12, ws.max_row + 1)):  # Skip header row
                    cell_value = ws.cell(row=row, column=excel_col_num).value
                    if cell_value is not None:
                        sample_values.append((row, cell_value))
                
                if sample_values:
                    print(f"  Sample values:")
                    for row_num, val in sample_values[:5]:
                        print(f"    Row {row_num}: {val} (type: {type(val).__name__})")
                else:
                    print(f"  No non-null values found in first 10 rows")
                    
            else:
                excel_letter = chr(65 + excel_col_num - 1) if excel_col_num <= 26 else chr(65 + (excel_col_num-1)//26 - 1) + chr(65 + (excel_col_num-1)%26)
                print(f"\nColumn {excel_col_num:2d} ({excel_letter}): NOT FOUND - File only has {ws.max_column} columns")
        
        wb.close()
        return headers, target_columns
        
    except Exception as e:
        print(f"Error reading Excel file with openpyxl: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def analyze_with_pandas():
    """
    Analyze Excel file using pandas
    """
    excel_file = "BFG-CO2H-HEX.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} not found in current directory")
        return None, None
    
    try:
        # Read the Excel file
        print(f"Reading Excel file with pandas: {excel_file}")
        df = pd.read_excel(excel_file, sheet_name=0)  # Read first sheet
        
        print(f"\n=== EXCEL FILE ANALYSIS (pandas) ===")
        print(f"File: {excel_file}")
        print(f"Sheet dimensions: {df.shape[0]} rows x {df.shape[1]} columns")
        
        # Show all column headers with their positions
        print(f"\n=== ALL COLUMN HEADERS ===")
        for i, col in enumerate(df.columns):
            excel_col_letter = chr(65 + i) if i < 26 else chr(65 + i//26 - 1) + chr(65 + i%26)
            print(f"Column {i+1:2d} ({excel_col_letter:2s}): {col}")
        
        # Focus on columns I through N (positions 8-13 in 0-based indexing)
        print(f"\n=== COLUMNS I THROUGH N ANALYSIS ===")
        target_columns = []
        target_positions = range(8, 14)  # Positions 8-13 (I-N in Excel)
        
        for pos in target_positions:
            if pos < len(df.columns):
                excel_letter = chr(65 + pos) if pos < 26 else chr(65 + pos//26 - 1) + chr(65 + pos%26)
                col_name = df.columns[pos]
                target_columns.append((pos, excel_letter, col_name))
                
                print(f"\nColumn {pos+1:2d} ({excel_letter}): '{col_name}'")
                print(f"  Data type: {df.dtypes[pos]}")
                print(f"  Non-null count: {df[col_name].count()}/{len(df)}")
                print(f"  Unique values: {df[col_name].nunique()}")
                
                # Show sample data (first 5 non-null values)
                sample_data = df[col_name].dropna().head(5)
                if len(sample_data) > 0:
                    print(f"  Sample values:")
                    for idx, val in sample_data.items():
                        print(f"    Row {idx+1}: {val}")
                else:
                    print(f"  No non-null values found")
                    
                # Show some statistics for numeric columns
                if df[col_name].dtype in ['int64', 'float64']:
                    stats = df[col_name].describe()
                    print(f"  Statistics:")
                    print(f"    Min: {stats['min']:.2f}")
                    print(f"    Max: {stats['max']:.2f}")
                    print(f"    Mean: {stats['mean']:.2f}")
            else:
                excel_letter = chr(65 + pos) if pos < 26 else chr(65 + pos//26 - 1) + chr(65 + pos%26)
                print(f"\nColumn {pos+1:2d} ({excel_letter}): NOT FOUND - File only has {len(df.columns)} columns")
        
        # Show a sample of the data for columns I-N
        print(f"\n=== SAMPLE DATA FOR COLUMNS I-N ===")
        if len(target_columns) > 0:
            sample_cols = [col_name for _, _, col_name in target_columns]
            sample_df = df[sample_cols].head(10)
            print(sample_df.to_string())
        
        # Check for missing data patterns
        print(f"\n=== MISSING DATA ANALYSIS ===")
        for pos, excel_letter, col_name in target_columns:
            missing_count = df[col_name].isnull().sum()
            missing_pct = (missing_count / len(df)) * 100
            print(f"Column {excel_letter} ({col_name}): {missing_count} missing ({missing_pct:.1f}%)")
        
        # Create column mapping summary
        print(f"\n=== COLUMN MAPPING SUMMARY ===")
        print("Excel Column -> Python Index -> Column Name")
        for pos, excel_letter, col_name in target_columns:
            print(f"{excel_letter:2s}           -> {pos:2d}            -> {col_name}")
        
        return df, target_columns
        
    except Exception as e:
        print(f"Error reading Excel file with pandas: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def analyze_excel_columns():
    """
    Main analysis function that tries different methods
    """
    print("Available libraries:")
    print(f"  pandas: {PANDAS_AVAILABLE}")
    print(f"  openpyxl: {OPENPYXL_AVAILABLE}")
    
    if PANDAS_AVAILABLE:
        print("\nTrying pandas method...")
        return analyze_with_pandas()
    elif OPENPYXL_AVAILABLE:
        print("\nTrying openpyxl method...")
        return analyze_with_openpyxl()
    else:
        print("\nNeither pandas nor openpyxl available. Please install one of them.")
        print("pip install pandas openpyxl")
        return None, None

def main():
    """Main function to run the analysis"""
    print("Excel Column Analyzer for Heat Exchanger Data")
    print("=" * 50)
    
    df, target_columns = analyze_excel_columns()
    
    if df is not None:
        print(f"\n=== ANALYSIS COMPLETE ===")
        print(f"Successfully analyzed {len(df)} rows of data")
        print(f"Found {len(target_columns)} target columns (I-N)")
        
        # Save column info to a text file for reference
        with open("column_analysis_results.txt", "w", encoding="utf-8") as f:
            f.write("Excel Column Analysis Results\n")
            f.write("=" * 30 + "\n\n")
            f.write(f"File: BFG-CO2H-HEX.xlsx\n")
            f.write(f"Dimensions: {df.shape[0]} rows x {df.shape[1]} columns\n\n")
            
            f.write("Columns I-N Mapping:\n")
            for pos, excel_letter, col_name in target_columns:
                f.write(f"  {excel_letter} -> Index {pos} -> '{col_name}'\n")
        
        print("Results saved to: column_analysis_results.txt")
    else:
        print("Analysis failed - please check the file path and format")

if __name__ == "__main__":
    main()