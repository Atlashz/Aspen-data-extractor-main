#!/usr/bin/env python3
"""
Economic Excel Exporter

生成详细的经济分析Excel报告，包含：
- 项目概览和财务指标
- CAPEX和OPEX详细分解
- 设备成本明细和参数
- 计算逻辑和假设条件
- 敏感性分析和图表

Author: TEA Analysis Framework
Date: 2025-07-27
Version: 1.0
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
import logging

# Excel processing
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.chart import BarChart, PieChart, Reference, LineChart
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Data processing
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Import local data structures
from data_interfaces import (
    EconomicAnalysisResults, CostItem, CapexData, OpexData, 
    FinancialParameters, EquipmentSizeData, CostCategory, CurrencyType
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class EconomicExcelExporter:
    """
    Excel报告生成器主类
    
    将经济分析结果导出为格式化的Excel文件，包含多个工作表
    和专业的图表、数据透视表等可视化元素。
    """
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        self.wb = None
        self.styles_initialized = False
        self.chart_count = 0
        
        # Define standard colors for charts and tables
        self.colors = {
            'primary': '4472C4',      # Blue
            'secondary': '70AD47',    # Green
            'accent1': 'FFC000',      # Orange
            'accent2': 'C5504B',      # Red
            'accent3': '7030A0',      # Purple
            'light_gray': 'F2F2F2',   # Light gray
            'dark_gray': '595959'     # Dark gray
        }
    
    def export_economic_analysis(self, results: EconomicAnalysisResults, 
                                output_file: str) -> str:
        """
        导出完整的经济分析报告到Excel文件
        
        Args:
            results: 经济分析结果数据
            output_file: 输出Excel文件路径
            
        Returns:
            生成的Excel文件路径
        """
        logger.info(f"🔄 Generating economic analysis report: {output_file}")
        
        # Create new workbook
        self.wb = Workbook()
        
        # Initialize styles
        self._initialize_styles()
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        try:
            # Create all worksheets
            self._create_executive_summary(results)
            self._create_capex_breakdown(results)
            self._create_opex_analysis(results)
            self._create_equipment_details(results)
            self._create_financial_analysis(results)
            self._create_sensitivity_analysis(results)
            self._create_calculation_parameters(results)
            self._create_assumptions_notes(results)
            
            # Save workbook
            output_path = Path(output_file)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            self.wb.save(output_file)
            
            logger.info(f"✅ Economic analysis report saved: {output_file}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise
    
    def _initialize_styles(self):
        """初始化Excel样式"""
        if self.styles_initialized:
            return
        
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, size=14, color='FFFFFF')
        header_style.fill = PatternFill(start_color=self.colors['primary'], 
                                       end_color=self.colors['primary'], 
                                       fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Subheader style
        subheader_style = NamedStyle(name="subheader")
        subheader_style.font = Font(bold=True, size=12)
        subheader_style.fill = PatternFill(start_color=self.colors['light_gray'], 
                                         end_color=self.colors['light_gray'], 
                                         fill_type='solid')
        subheader_style.alignment = Alignment(horizontal='left', vertical='center')
        
        # Currency style
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = '"$"#,##0'
        currency_style.alignment = Alignment(horizontal='right')
        
        # Percentage style
        percentage_style = NamedStyle(name="percentage")
        percentage_style.number_format = '0.0%'
        percentage_style.alignment = Alignment(horizontal='right')
        
        # Add styles to workbook
        try:
            existing_styles = []
            try:
                existing_styles = [getattr(style, 'name', str(style)) for style in self.wb.named_styles]
            except:
                existing_styles = []
            
            if "header" not in existing_styles:
                self.wb.add_named_style(header_style)
            if "subheader" not in existing_styles:
                self.wb.add_named_style(subheader_style)
            if "currency" not in existing_styles:
                self.wb.add_named_style(currency_style)
            if "percentage" not in existing_styles:
                self.wb.add_named_style(percentage_style)
        except ValueError:
            # Styles already exist
            pass
        
        self.styles_initialized = True
    
    def _create_executive_summary(self, results: EconomicAnalysisResults):
        """创建项目概览工作表"""
        ws = self.wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = f"Economic Analysis Report - {results.project_name}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')
        
        # Project information
        row = 3
        ws[f'A{row}'] = "Project Information"
        ws[f'A{row}'].style = 'subheader'
        
        project_info = [
            ("Project Name", results.project_name),
            ("Analysis Date", results.timestamp.strftime("%Y-%m-%d %H:%M")),
            ("Analysis Version", results.analysis_version),
            ("Confidence Level", results.confidence_level or "Medium"),
            ("Accuracy Range", results.accuracy_range or "±25%")
        ]
        
        for i, (label, value) in enumerate(project_info):
            row_num = row + 1 + i
            ws[f'A{row_num}'] = label
            ws[f'B{row_num}'] = value
        
        # Financial summary
        row += len(project_info) + 3
        ws[f'A{row}'] = "Financial Summary"
        ws[f'A{row}'].style = 'subheader'
        
        financial_summary = [
            ("Total CAPEX", results.total_capex, "currency"),
            ("Annual OPEX", results.annual_opex, "currency"),
            ("Production Cost", results.production_cost, "currency"),
            ("Net Present Value (NPV)", results.npv, "currency"),
            ("Internal Rate of Return (IRR)", results.irr, "percentage"),
            ("Payback Period", f"{results.payback_period:.1f} years" if results.payback_period else "N/A", None)
        ]
        
        for i, (label, value, style) in enumerate(financial_summary):
            row_num = row + 1 + i
            ws[f'A{row_num}'] = label
            cell = ws[f'B{row_num}']
            if isinstance(value, (int, float)) and style:
                cell.value = value
                cell.style = style
            else:
                cell.value = value
        
        # Equipment summary
        row += len(financial_summary) + 3
        ws[f'A{row}'] = "Equipment Summary"
        ws[f'A{row}'].style = 'subheader'
        
        equipment_count = len(results.equipment_list)
        ws[f'A{row+1}'] = "Total Equipment Count"
        ws[f'B{row+1}'] = equipment_count
        
        # Data sources
        row += 4
        ws[f'A{row}'] = "Data Sources"
        ws[f'A{row}'].style = 'subheader'
        
        for i, source in enumerate(results.data_sources):
            ws[f'A{row+1+i}'] = f"• {source}"
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Add CAPEX/OPEX pie chart
        self._add_capex_opex_pie_chart(ws, results)
    
    def _create_capex_breakdown(self, results: EconomicAnalysisResults):
        """创建CAPEX分解工作表"""
        ws = self.wb.create_sheet("CAPEX Breakdown")
        
        # Title
        ws['A1'] = "Capital Expenditure (CAPEX) Breakdown"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:H1')
        
        # Equipment costs table
        row = 3
        ws[f'A{row}'] = "Equipment Costs"
        ws[f'A{row}'].style = 'subheader'
        
        # Headers
        headers = ["Equipment Name", "Category", "Base Cost", "Quantity", "Installation Factor", 
                  "Material Factor", "Location Factor", "Installed Cost"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=row+1, column=i+1, value=header)
            cell.style = 'subheader'
        
        # Equipment data
        equipment_data = []
        if results.capex_data and results.capex_data.equipment_costs:
            for item in results.capex_data.equipment_costs.values():
                try:
                    equipment_data.append([
                        getattr(item, 'name', str(item)),
                        getattr(item.category, 'value', str(item.category)) if hasattr(item, 'category') else 'Unknown',
                        getattr(item, 'base_cost', 0.0),
                        getattr(item, 'quantity', 1.0),
                        getattr(item, 'installation_factor', 1.0),
                        getattr(item, 'material_factor', 1.0),
                        getattr(item, 'location_factor', 1.0),
                        item.calculate_installed_cost() if hasattr(item, 'calculate_installed_cost') else 0.0
                    ])
                except Exception as e:
                    # Handle case where item might be a string or other unexpected type
                    equipment_data.append([
                        str(item),
                        'Unknown',
                        0.0,
                        1.0,
                        1.0,
                        1.0,
                        1.0,
                        0.0
                    ])
        
        # Add equipment data to worksheet
        for i, row_data in enumerate(equipment_data):
            row_num = row + 2 + i
            for j, value in enumerate(row_data):
                cell = ws.cell(row=row_num, column=j+1, value=value)
                if j in [2, 7]:  # Cost columns
                    cell.style = 'currency'
        
        # Installation costs table
        row += len(equipment_data) + 4
        ws[f'A{row}'] = "Installation & Indirect Costs"
        ws[f'A{row}'].style = 'subheader'
        
        # Installation costs headers
        install_headers = ["Cost Item", "Category", "Base Cost", "Method", "Total Cost"]
        for i, header in enumerate(install_headers):
            cell = ws.cell(row=row+1, column=i+1, value=header)
            cell.style = 'subheader'
        
        # Installation costs data
        install_costs = list(results.capex_data.installation_costs.values()) + \
                       list(results.capex_data.indirect_costs.values())
        
        for i, item in enumerate(install_costs):
            row_num = row + 2 + i
            ws.cell(row=row_num, column=1, value=getattr(item, 'name', str(item)))
            ws.cell(row=row_num, column=2, value=item.category.value)
            ws.cell(row=row_num, column=3, value=item.base_cost).style = 'currency'
            ws.cell(row=row_num, column=4, value=item.estimation_method or "Standard")
            ws.cell(row=row_num, column=5, value=item.calculate_installed_cost()).style = 'currency'
        
        # Total CAPEX summary
        row += len(install_costs) + 3
        ws[f'A{row}'] = "CAPEX Summary"
        ws[f'A{row}'].style = 'subheader'
        
        capex_summary = [
            ("Equipment Subtotal", sum(item.calculate_installed_cost() 
                                     for item in results.capex_data.equipment_costs.values())),
            ("Installation Subtotal", sum(item.calculate_installed_cost() 
                                        for item in results.capex_data.installation_costs.values())),
            ("Indirect Costs", sum(item.calculate_installed_cost() 
                                 for item in results.capex_data.indirect_costs.values())),
            ("Contingency", results.total_capex * results.capex_data.contingency_rate),
            ("Total CAPEX", results.total_capex)
        ]
        
        for i, (label, value) in enumerate(capex_summary):
            row_num = row + 1 + i
            ws[f'A{row_num}'] = label
            ws[f'B{row_num}'].value = value
            ws[f'B{row_num}'].style = 'currency'
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
        # Add CAPEX breakdown bar chart
        self._add_capex_breakdown_chart(ws, results, row + len(capex_summary) + 2)
    
    def _create_opex_analysis(self, results: EconomicAnalysisResults):
        """创建OPEX分析工作表"""
        ws = self.wb.create_sheet("OPEX Analysis")
        
        # Title
        ws['A1'] = "Operating Expenditure (OPEX) Analysis"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')
        
        # Raw materials table
        row = 3
        if results.opex_data.raw_material_costs:
            ws[f'A{row}'] = "Raw Material Costs"
            ws[f'A{row}'].style = 'subheader'
            
            self._create_opex_table(ws, row, results.opex_data.raw_material_costs.values(), 
                                  "Raw Materials")
            row += len(results.opex_data.raw_material_costs) + 4
        
        # Utility costs table
        if results.opex_data.utility_costs:
            ws[f'A{row}'] = "Utility Costs"
            ws[f'A{row}'].style = 'subheader'
            
            self._create_opex_table(ws, row, results.opex_data.utility_costs.values(), 
                                  "Utilities")
            row += len(results.opex_data.utility_costs) + 4
        
        # Labor costs table
        if results.opex_data.labor_costs:
            ws[f'A{row}'] = "Labor Costs"
            ws[f'A{row}'].style = 'subheader'
            
            self._create_opex_table(ws, row, results.opex_data.labor_costs.values(), 
                                  "Labor")
            row += len(results.opex_data.labor_costs) + 4
        
        # Maintenance costs table
        if results.opex_data.maintenance_costs:
            ws[f'A{row}'] = "Maintenance Costs"
            ws[f'A{row}'].style = 'subheader'
            
            self._create_opex_table(ws, row, results.opex_data.maintenance_costs.values(), 
                                  "Maintenance")
            row += len(results.opex_data.maintenance_costs) + 4
        
        # OPEX summary
        ws[f'A{row}'] = "Annual OPEX Summary"
        ws[f'A{row}'].style = 'subheader'
        
        opex_summary = [
            ("Raw Materials", sum(item.calculate_installed_cost() 
                                for item in results.opex_data.raw_material_costs.values())),
            ("Utilities", sum(item.calculate_installed_cost() 
                            for item in results.opex_data.utility_costs.values())),
            ("Labor", sum(item.calculate_installed_cost() 
                        for item in results.opex_data.labor_costs.values())),
            ("Maintenance", sum(item.calculate_installed_cost() 
                              for item in results.opex_data.maintenance_costs.values())),
            ("Total Annual OPEX", results.annual_opex)
        ]
        
        for i, (label, value) in enumerate(opex_summary):
            row_num = row + 1 + i
            ws[f'A{row_num}'] = label
            ws[f'B{row_num}'].value = value
            ws[f'B{row_num}'].style = 'currency'
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
        # Add OPEX breakdown chart
        self._add_opex_breakdown_chart(ws, results, row + len(opex_summary) + 2)
    
    def _create_equipment_details(self, results: EconomicAnalysisResults):
        """创建设备详细信息工作表"""
        ws = self.wb.create_sheet("Equipment Details")
        
        # Title
        ws['A1'] = "Equipment Sizing and Costing Details"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:K1')
        
        # Equipment table headers
        row = 3
        headers = ["Equipment Name", "Type", "Volume (m³)", "Area (m²)", "Diameter (m)", 
                  "Height (m)", "Power (kW)", "Design P (bar)", "Design T (°C)", 
                  "Estimated Cost", "Cost Basis"]
        
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=header)
            cell.style = 'subheader'
        
        # Equipment data
        for i, (name, equipment) in enumerate(results.equipment_list.items()):
            row_num = row + 1 + i
            
            data = [
                getattr(equipment, 'name', str(equipment)),
                equipment.equipment_type.value,
                equipment.volume,
                equipment.area,
                equipment.diameter,
                equipment.height,
                equipment.power_rating,
                equipment.design_pressure,
                equipment.design_temperature,
                equipment.estimated_cost,
                equipment.cost_basis or "2024 USD"
            ]
            
            for j, value in enumerate(data):
                cell = ws.cell(row=row_num, column=j+1)
                if value is not None:
                    cell.value = value
                    if j == 9:  # Cost column
                        cell.style = 'currency'
                else:
                    cell.value = "N/A"
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_financial_analysis(self, results: EconomicAnalysisResults):
        """创建财务分析工作表"""
        ws = self.wb.create_sheet("Financial Analysis")
        
        # Title
        ws['A1'] = "Financial Analysis and Cash Flow"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:F1')
        
        # Financial parameters
        row = 3
        ws[f'A{row}'] = "Financial Parameters"
        ws[f'A{row}'].style = 'subheader'
        
        params = results.financial_params
        financial_params = [
            ("Project Life", f"{params.project_life} years"),
            ("Discount Rate", params.discount_rate),
            ("Tax Rate", params.tax_rate),
            ("Depreciation Method", params.depreciation_method),
            ("Depreciation Life", f"{params.depreciation_life} years"),
            ("Annual Production", f"{params.annual_production:,.0f} kg/year" if params.annual_production else "N/A")
        ]
        
        for i, (label, value) in enumerate(financial_params):
            row_num = row + 1 + i
            ws[f'A{row_num}'] = label
            cell = ws[f'B{row_num}']
            if isinstance(value, float) and 0 < value < 1:
                cell.value = value
                cell.style = 'percentage'
            else:
                cell.value = value
        
        # Economic indicators
        row += len(financial_params) + 3
        ws[f'A{row}'] = "Economic Indicators"
        ws[f'A{row}'].style = 'subheader'
        
        indicators = [
            ("Net Present Value (NPV)", results.npv, "currency"),
            ("Internal Rate of Return (IRR)", results.irr, "percentage"),
            ("Payback Period", f"{results.payback_period:.1f} years" if results.payback_period else "N/A", None),
            ("Production Cost", results.production_cost, "currency"),
            ("Break-even Price", results.break_even_price, "currency")
        ]
        
        for i, (label, value, style) in enumerate(indicators):
            row_num = row + 1 + i
            ws[f'A{row_num}'] = label
            cell = ws[f'B{row_num}']
            if isinstance(value, (int, float)) and style:
                cell.value = value
                cell.style = style
            else:
                cell.value = value
        
        # Cash flow analysis (if available)
        if results.financial_params.annual_cash_flows:
            row += len(indicators) + 3
            ws[f'A{row}'] = "Cash Flow Analysis"
            ws[f'A{row}'].style = 'subheader'
            
            # Cash flow table headers
            cf_headers = ["Year", "Cash Flow", "Cumulative Cash Flow"]
            for i, header in enumerate(cf_headers):
                cell = ws.cell(row=row+1, column=i+1, value=header)
                cell.style = 'subheader'
            
            # Cash flow data
            cash_flows = results.financial_params.annual_cash_flows
            cumulative = 0
            
            for i, cf in enumerate(cash_flows[:21]):  # Limit to 21 years (initial + 20 years)
                row_num = row + 2 + i
                cumulative += cf
                
                ws.cell(row=row_num, column=1, value=i)
                ws.cell(row=row_num, column=2, value=cf).style = 'currency'
                ws.cell(row=row_num, column=3, value=cumulative).style = 'currency'
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_sensitivity_analysis(self, results: EconomicAnalysisResults):
        """创建敏感性分析工作表"""
        ws = self.wb.create_sheet("Sensitivity Analysis")
        
        # Title
        ws['A1'] = "Sensitivity Analysis"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:F1')
        
        # Note about sensitivity analysis
        row = 3
        ws[f'A{row}'] = "Parameter Sensitivity Analysis"
        ws[f'A{row}'].style = 'subheader'
        
        # Create a sample sensitivity analysis
        sensitive_params = [
            ("CAPEX", [-30, -20, -10, 0, 10, 20, 30]),
            ("OPEX", [-30, -20, -10, 0, 10, 20, 30]),
            ("Discount Rate", [-3, -2, -1, 0, 1, 2, 3]),
            ("Product Price", [-30, -20, -10, 0, 10, 20, 30])
        ]
        
        # Headers
        row += 2
        headers = ["Parameter", "-30%", "-20%", "-10%", "Base", "+10%", "+20%", "+30%"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=header)
            cell.style = 'subheader'
        
        # Calculate sensitivity for NPV
        base_npv = results.npv
        
        for param_idx, (param_name, variations) in enumerate(sensitive_params):
            row_num = row + 1 + param_idx
            ws.cell(row=row_num, column=1, value=param_name)
            
            for var_idx, variation in enumerate(variations):
                # Simple sensitivity calculation (simplified)
                if param_name == "CAPEX":
                    sensitive_npv = base_npv - (results.total_capex * variation / 100)
                elif param_name == "OPEX":
                    # Simplified: assume OPEX affects NPV over project life
                    npv_change = (results.annual_opex * variation / 100) * results.financial_params.project_life
                    sensitive_npv = base_npv - npv_change
                else:
                    # Default sensitivity
                    sensitive_npv = base_npv * (1 + variation / 100)
                
                cell = ws.cell(row=row_num, column=var_idx+2, value=sensitive_npv)
                cell.style = 'currency'
                
                # Color coding for negative values
                if sensitive_npv < 0:
                    cell.fill = PatternFill(start_color=self.colors['accent2'], 
                                          end_color=self.colors['accent2'], 
                                          fill_type='solid')
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_calculation_parameters(self, results: EconomicAnalysisResults):
        """创建计算参数工作表"""
        ws = self.wb.create_sheet("Calculation Parameters")
        
        # Title
        ws['A1'] = "Calculation Parameters and Methods"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:D1')
        
        # Estimation methods
        row = 3
        ws[f'A{row}'] = "Estimation Methods Used"
        ws[f'A{row}'].style = 'subheader'
        
        for i, method in enumerate(results.estimation_methods):
            ws[f'A{row+1+i}'] = f"• {method}"
        
        # Cost factors and assumptions
        row += len(results.estimation_methods) + 3
        ws[f'A{row}'] = "Standard Cost Factors"
        ws[f'A{row}'].style = 'subheader'
        
        cost_factors = [
            ("Installation Factor", "2.5", "Typical for process equipment"),
            ("Engineering & Design", "12%", "Percentage of equipment cost"),
            ("Construction Management", "8%", "Percentage of equipment cost"),
            ("Contingency", f"{results.capex_data.contingency_rate*100:.0f}%", "Project contingency"),
            ("Maintenance Rate", f"{results.opex_data.maintenance_rate*100:.0f}%", "Annual percentage of CAPEX"),
            ("Labor Rate", "2%", "Annual percentage of CAPEX")
        ]
        
        # Headers
        factor_headers = ["Parameter", "Value", "Description"]
        for i, header in enumerate(factor_headers):
            cell = ws.cell(row=row+1, column=i+1, value=header)
            cell.style = 'subheader'
        
        # Cost factors data
        for i, (param, value, desc) in enumerate(cost_factors):
            row_num = row + 2 + i
            ws.cell(row=row_num, column=1, value=param)
            ws.cell(row=row_num, column=2, value=value)
            ws.cell(row=row_num, column=3, value=desc)
        
        # Equipment costing correlations
        row += len(cost_factors) + 4
        ws[f'A{row}'] = "Equipment Costing Correlations"
        ws[f'A{row}'].style = 'subheader'
        
        correlations = [
            ("Reactor", "Cost = $50,000 × (Volume_m³)^0.6"),
            ("Pump", "Cost = $5,000 × (Power_kW)^0.7"),
            ("Compressor", "Cost = $15,000 × (Power_kW)^0.7"),
            ("Heat Exchanger", "Cost = $1,000 × (Area_m²)^0.65"),
            ("Distillation Column", "Cost = $25,000 × (Diameter_m)^1.5 × (Height_m)^0.8"),
            ("Separator", "Cost = $20,000 × (Volume_m³)^0.6"),
            ("Tank/Vessel", "Cost = $8,000 × (Volume_m³)^0.7")
        ]
        
        corr_headers = ["Equipment Type", "Cost Correlation"]
        for i, header in enumerate(corr_headers):
            cell = ws.cell(row=row+1, column=i+1, value=header)
            cell.style = 'subheader'
        
        for i, (eq_type, correlation) in enumerate(correlations):
            row_num = row + 2 + i
            ws.cell(row=row_num, column=1, value=eq_type)
            ws.cell(row=row_num, column=2, value=correlation)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_assumptions_notes(self, results: EconomicAnalysisResults):
        """创建假设和备注工作表"""
        ws = self.wb.create_sheet("Assumptions & Notes")
        
        # Title
        ws['A1'] = "Assumptions and Notes"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:D1')
        
        # Key assumptions
        row = 3
        ws[f'A{row}'] = "Key Assumptions"
        ws[f'A{row}'].style = 'subheader'
        
        default_assumptions = [
            "All costs are in 2024 USD",
            "Plant operates 8760 hours per year (100% availability)",
            "Installation factors are typical for process industry",
            "Utility prices are based on industrial averages",
            "Equipment costs based on carbon steel construction",
            "Location factor = 1.0 (US Gulf Coast basis)",
            "Straight-line depreciation over 10 years",
            "Corporate tax rate = 25%",
            "Discount rate = 10%"
        ]
        
        all_assumptions = default_assumptions + results.assumptions
        
        for i, assumption in enumerate(all_assumptions):
            ws[f'A{row+1+i}'] = f"• {assumption}"
        
        # Data sources
        row += len(all_assumptions) + 3
        ws[f'A{row}'] = "Data Sources"
        ws[f'A{row}'].style = 'subheader'
        
        for i, source in enumerate(results.data_sources):
            ws[f'A{row+1+i}'] = f"• {source}"
        
        # Analysis metadata
        row += len(results.data_sources) + 3
        ws[f'A{row}'] = "Analysis Metadata"
        ws[f'A{row}'].style = 'subheader'
        
        metadata = [
            ("Analysis Date", results.timestamp.strftime("%Y-%m-%d %H:%M:%S")),
            ("Analysis Version", results.analysis_version),
            ("Confidence Level", results.confidence_level or "Medium"),
            ("Accuracy Range", results.accuracy_range or "±25%"),
            ("Equipment Count", len(results.equipment_list)),
            ("Data Sources Count", len(results.data_sources))
        ]
        
        for i, (label, value) in enumerate(metadata):
            row_num = row + 1 + i
            ws[f'A{row_num}'] = label
            ws[f'B{row_num}'] = value
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_opex_table(self, ws, start_row, cost_items, table_name):
        """创建OPEX数据表"""
        headers = ["Item Name", "Category", "Annual Cost", "Unit", "Quantity", "Method"]
        
        # Headers
        for i, header in enumerate(headers):
            cell = ws.cell(row=start_row+1, column=i+1, value=header)
            cell.style = 'subheader'
        
        # Data
        for i, item in enumerate(cost_items):
            row_num = start_row + 2 + i
            ws.cell(row=row_num, column=1, value=getattr(item, 'name', str(item)))
            ws.cell(row=row_num, column=2, value=item.category.value)
            ws.cell(row=row_num, column=3, value=item.calculate_installed_cost()).style = 'currency'
            ws.cell(row=row_num, column=4, value=item.unit)
            ws.cell(row=row_num, column=5, value=item.quantity)
            ws.cell(row=row_num, column=6, value=item.estimation_method or "Standard")
    
    def _auto_adjust_columns(self, ws):
        """自动调整列宽"""
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils import get_column_letter
        
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            # Find the first non-merged cell to get column letter
            for cell in column:
                if not isinstance(cell, MergedCell):
                    column_letter = get_column_letter(cell.column)
                    break
            
            if column_letter is None:
                continue
            
            for cell in column:
                try:
                    if not isinstance(cell, MergedCell) and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_capex_opex_pie_chart(self, ws, results: EconomicAnalysisResults):
        """添加CAPEX/OPEX饼图"""
        try:
            # Create pie chart data
            chart_data = [
                ["Category", "Amount"],
                ["CAPEX", results.total_capex],
                ["Annual OPEX", results.annual_opex]
            ]
            
            # Add data to worksheet (in a location that won't interfere)
            chart_start_row = 20
            for i, row_data in enumerate(chart_data):
                for j, value in enumerate(row_data):
                    ws.cell(row=chart_start_row + i, column=4 + j, value=value)
            
            # Create pie chart
            chart = PieChart()
            chart.title = "CAPEX vs Annual OPEX"
            
            # Data references
            data = Reference(ws, min_col=5, min_row=chart_start_row + 1, 
                           max_row=chart_start_row + len(chart_data) - 1)
            categories = Reference(ws, min_col=4, min_row=chart_start_row + 1, 
                                 max_row=chart_start_row + len(chart_data) - 1)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(categories)
            
            # Chart styling
            chart.width = 15
            chart.height = 10
            
            # Add chart to worksheet
            ws.add_chart(chart, "D3")
            
        except Exception as e:
            logger.warning(f"Could not create CAPEX/OPEX pie chart: {str(e)}")
    
    def _add_capex_breakdown_chart(self, ws, results: EconomicAnalysisResults, start_row):
        """添加CAPEX分解柱状图"""
        try:
            # Prepare chart data
            equipment_total = sum(item.calculate_installed_cost() 
                                for item in results.capex_data.equipment_costs.values())
            installation_total = sum(item.calculate_installed_cost() 
                                   for item in results.capex_data.installation_costs.values())
            indirect_total = sum(item.calculate_installed_cost() 
                               for item in results.capex_data.indirect_costs.values())
            
            chart_data = [
                ["Category", "Cost"],
                ["Equipment", equipment_total],
                ["Installation", installation_total],
                ["Indirect", indirect_total],
                ["Contingency", results.total_capex * results.capex_data.contingency_rate]
            ]
            
            # Add data to worksheet
            for i, row_data in enumerate(chart_data):
                for j, value in enumerate(row_data):
                    ws.cell(row=start_row + i, column=4 + j, value=value)
            
            # Create bar chart
            chart = BarChart()
            chart.title = "CAPEX Breakdown"
            chart.y_axis.title = "Cost (USD)"
            
            # Data references
            data = Reference(ws, min_col=5, min_row=start_row + 1, 
                           max_row=start_row + len(chart_data) - 1)
            categories = Reference(ws, min_col=4, min_row=start_row + 1, 
                                 max_row=start_row + len(chart_data) - 1)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(categories)
            
            # Chart styling
            chart.width = 15
            chart.height = 10
            
            # Add chart to worksheet
            ws.add_chart(chart, f"D{start_row + len(chart_data) + 1}")
            
        except Exception as e:
            logger.warning(f"Could not create CAPEX breakdown chart: {str(e)}")
    
    def _add_opex_breakdown_chart(self, ws, results: EconomicAnalysisResults, start_row):
        """添加OPEX分解柱状图"""
        try:
            # Prepare chart data
            raw_materials_total = sum(item.calculate_installed_cost() 
                                    for item in results.opex_data.raw_material_costs.values())
            utilities_total = sum(item.calculate_installed_cost() 
                                for item in results.opex_data.utility_costs.values())
            labor_total = sum(item.calculate_installed_cost() 
                            for item in results.opex_data.labor_costs.values())
            maintenance_total = sum(item.calculate_installed_cost() 
                                  for item in results.opex_data.maintenance_costs.values())
            
            chart_data = [
                ["Category", "Annual Cost"],
                ["Raw Materials", raw_materials_total],
                ["Utilities", utilities_total],
                ["Labor", labor_total],
                ["Maintenance", maintenance_total]
            ]
            
            # Filter out zero values
            chart_data = [chart_data[0]] + [row for row in chart_data[1:] if row[1] > 0]
            
            # Add data to worksheet
            for i, row_data in enumerate(chart_data):
                for j, value in enumerate(row_data):
                    ws.cell(row=start_row + i, column=4 + j, value=value)
            
            # Create bar chart
            chart = BarChart()
            chart.title = "Annual OPEX Breakdown"
            chart.y_axis.title = "Annual Cost (USD)"
            
            # Data references
            data = Reference(ws, min_col=5, min_row=start_row + 1, 
                           max_row=start_row + len(chart_data) - 1)
            categories = Reference(ws, min_col=4, min_row=start_row + 1, 
                                 max_row=start_row + len(chart_data) - 1)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(categories)
            
            # Chart styling
            chart.width = 15
            chart.height = 10
            
            # Add chart to worksheet
            ws.add_chart(chart, f"D{start_row + len(chart_data) + 1}")
            
        except Exception as e:
            logger.warning(f"Could not create OPEX breakdown chart: {str(e)}")


def test_excel_exporter():
    """测试Excel导出器"""
    try:
        # Create sample economic results for testing
        from datetime import datetime
        
        sample_results = EconomicAnalysisResults(
            project_name="Test Economic Analysis",
            timestamp=datetime.now()
        )
        
        # Add some sample data
        sample_results.total_capex = 5000000
        sample_results.annual_opex = 1200000
        sample_results.npv = 2500000
        sample_results.irr = 0.15
        sample_results.payback_period = 4.2
        
        # Create exporter and generate report
        exporter = EconomicExcelExporter()
        output_file = "test_economic_report.xlsx"
        
        result_file = exporter.export_economic_analysis(sample_results, output_file)
        print(f"✅ Test Excel report generated: {result_file}")
        
    except Exception as e:
        print(f"❌ Error testing Excel exporter: {str(e)}")


if __name__ == "__main__":
    test_excel_exporter()