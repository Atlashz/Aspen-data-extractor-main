#!/usr/bin/env python3
"""
Simple Excel reader using openpyxl to analyze columns I-N
"""

import sys
import os

try:
    from openpyxl import load_workbook
    print("openpyxl available")
except ImportError:
    print("openpyxl not available - trying to analyze with basic approach")
    sys.exit(1)

def analyze_excel_columns():
    excel_file = "BFG-CO2H-HEX.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} not found")
        return
    
    try:
        print(f"Opening {excel_file}...")
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        
        print(f"Active worksheet: {ws.title}")
        print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Get all headers (row 1)
        print("\n=== ALL COLUMN HEADERS ===")
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            excel_letter = chr(64 + col) if col <= 26 else chr(64 + col//26) + chr(64 + col%26)
            headers.append(header)
            print(f"Column {col:2d} ({excel_letter}): {header}")
        
        # Focus on columns I-N (9-14)
        print(f"\n=== COLUMNS I THROUGH N (positions 9-14) ===")
        for col_num in range(9, 15):  # I=9, J=10, K=11, L=12, M=13, N=14
            if col_num <= ws.max_column:
                excel_letter = chr(64 + col_num)
                header = ws.cell(row=1, column=col_num).value
                
                print(f"\nColumn {excel_letter} (position {col_num}):")
                print(f"  Header: '{header}'")
                
                # Get sample data from first 10 rows
                sample_data = []
                for row in range(2, min(12, ws.max_row + 1)):
                    cell_value = ws.cell(row=row, column=col_num).value
                    if cell_value is not None:
                        sample_data.append((row, cell_value, type(cell_value).__name__))
                
                if sample_data:
                    print(f"  Sample data:")
                    for row_num, value, data_type in sample_data[:5]:
                        print(f"    Row {row_num}: {value} ({data_type})")
                else:
                    print(f"  No data found in first 10 rows")
            else:
                excel_letter = chr(64 + col_num)
                print(f"\nColumn {excel_letter} (position {col_num}): NOT FOUND - only {ws.max_column} columns exist")
        
        # Summary mapping
        print(f"\n=== COLUMN MAPPING SUMMARY ===")
        print("Excel Letter -> Column Number -> Header Name")
        for col_num in range(9, min(15, ws.max_column + 1)):
            excel_letter = chr(64 + col_num)
            header = ws.cell(row=1, column=col_num).value
            print(f"{excel_letter:2s}           -> {col_num:2d}            -> '{header}'")
        
        wb.close()
        print(f"\n✅ Analysis complete!")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_excel_columns()